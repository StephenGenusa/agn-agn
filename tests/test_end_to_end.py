import io
import pathlib
import zipfile

import pytest
from openpyxl import load_workbook

from callsigns.cli import main
from callsigns.exporters.dta import decode_master_dta
from callsigns.store import WorkbookStore

FIXTURE = pathlib.Path(__file__).parent / "fixtures" / "pota_hunters.json"


@pytest.fixture
def fake_api(monkeypatch):
    raw = FIXTURE.read_bytes()

    class FakeClient:
        def get_bytes(self, url, **kwargs):
            return raw

    monkeypatch.setattr("callsigns.providers.pota.HttpClient", lambda: FakeClient())


def test_refresh_then_every_export(tmp_path, fake_api):
    store = tmp_path / "POTA-Hunters.xlsx"
    assert main(["refresh", "pota-hunters", "-y", "all", "--store", str(store)]) == 0
    for fmt, expected in [
        ("dta", "POTA_Calls_CW.dta"),
        ("scp", "POTA_Calls_CW.scp"),
        ("lst", "POTA_Calls_CW.lst"),
    ]:
        code = main(
            [
                "export",
                "pota-hunters",
                fmt,
                "-o",
                "cw",
                "--store",
                str(store),
                "--out",
                str(tmp_path),
            ]
        )
        assert code == 0
        assert (tmp_path / expected).is_file()

    calls = decode_master_dta((tmp_path / "POTA_Calls_CW.dta").read_bytes())
    assert "K2UPD" in calls
    assert all(call.isupper() for call in calls)


def test_xlsx_export_end_to_end(tmp_path, fake_api):
    store = tmp_path / "S.xlsx"
    main(["refresh", "pota-hunters", "-y", "all", "--store", str(store)])
    assert (
        main(
            [
                "export",
                "pota-hunters",
                "xlsx",
                "-o",
                "cw",
                "--store",
                str(store),
                "--out",
                str(tmp_path),
                "--basename",
                "export",
            ]
        )
        == 0
    )
    sheet = load_workbook(tmp_path / "export.xlsx").active
    assert sheet.max_row == 5  # header + 4 CW-active hunters
    assert sheet["A2"].value == "SM3NRY"


def test_refresh_is_idempotent(tmp_path, fake_api):
    store = tmp_path / "S.xlsx"
    main(["refresh", "pota-hunters", "-y", "all", "--store", str(store)])
    main(["refresh", "pota-hunters", "-y", "all", "--store", str(store)])
    assert WorkbookStore(store).sheet_names() == ["All-Time"]
    assert len(WorkbookStore(store).meta_rows()) == 1


def test_second_refresh_adds_a_sheet_without_losing_the_first(tmp_path, fake_api):
    store = tmp_path / "S.xlsx"
    main(["refresh", "pota-hunters", "-y", "all", "--store", str(store)])
    main(["refresh", "pota-hunters", "-y", "2026", "--store", str(store)])
    assert sorted(WorkbookStore(store).sheet_names()) == ["2026", "All-Time"]
    assert len(WorkbookStore(store).meta_rows()) == 2


def test_exports_from_different_periods_do_not_collide(tmp_path, fake_api):
    store = tmp_path / "S.xlsx"
    main(["refresh", "pota-hunters", "-y", "all,2026", "--store", str(store)])
    for period in ("all", "2026"):
        main(
            [
                "export",
                "pota-hunters",
                "scp",
                "-y",
                period,
                "-o",
                "cw",
                "--store",
                str(store),
                "--out",
                str(tmp_path),
            ]
        )
    assert (tmp_path / "POTA_Calls_CW.scp").is_file()
    assert (tmp_path / "POTA_Calls_CW_2026.scp").is_file()


@pytest.fixture
def fake_rbn(monkeypatch):
    sample = pathlib.Path(__file__).parent / "fixtures" / "rbn_sample.csv"

    class FakeClient:
        def get_bytes(self, url, **kwargs):
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, "w") as zf:
                zf.writestr("spots.csv", sample.read_text())
            return buffer.getvalue()

    monkeypatch.setattr("callsigns.cache.HttpClient", lambda: FakeClient())


def test_rbn_refresh_then_every_export(tmp_path, fake_rbn):
    store = tmp_path / "RBN-CW.xlsx"
    assert (
        main(["refresh", "rbn-cw", "-y", "20251129", "-o", "cw", "--store", str(store)])
        == 0
    )
    for fmt, expected in [
        ("dta", "RBN_Calls_CW_20251129.dta"),
        ("scp", "RBN_Calls_CW_20251129.scp"),
        ("lst", "RBN_Calls_CW_20251129.lst"),
    ]:
        assert (
            main(
                [
                    "export",
                    "rbn-cw",
                    fmt,
                    "-y",
                    "20251129",
                    "-o",
                    "cw",
                    "--store",
                    str(store),
                    "--out",
                    str(tmp_path),
                ]
            )
            == 0
        )
        assert (tmp_path / expected).is_file()
    calls = decode_master_dta((tmp_path / "RBN_Calls_CW_20251129.dta").read_bytes())
    assert "S59L" in calls


def test_rbn_ranks_by_spot_count(tmp_path, fake_rbn):
    store = tmp_path / "R.xlsx"
    main(["refresh", "rbn-cw", "-y", "20251129", "-o", "cw", "--store", str(store)])
    rows = WorkbookStore(store).read_sheet("20251129 CW")
    counts = [row["Spots"] for row in rows]
    assert counts == sorted(counts, reverse=True)


def test_rbn_xlsx_export_carries_wpm_columns(tmp_path, fake_rbn):
    store = tmp_path / "R.xlsx"
    main(["refresh", "rbn-cw", "-y", "20251129", "-o", "cw", "--store", str(store)])
    main(
        [
            "export",
            "rbn-cw",
            "xlsx",
            "-y",
            "20251129",
            "-o",
            "cw",
            "--store",
            str(store),
            "--out",
            str(tmp_path),
            "--basename",
            "rbn",
        ]
    )
    sheet = load_workbook(tmp_path / "rbn.xlsx").active
    headers = [c.value for c in sheet[1]]
    assert "WPM Median" in headers
    assert "Spots" in headers


def test_rbn_and_pota_stores_coexist(tmp_path, fake_api, fake_rbn):
    main(["refresh", "pota-hunters", "-y", "all", "--store", str(tmp_path / "P.xlsx")])
    main(
        [
            "refresh",
            "rbn-cw",
            "-y",
            "20251129",
            "-o",
            "cw",
            "--store",
            str(tmp_path / "R.xlsx"),
        ]
    )
    assert WorkbookStore(tmp_path / "P.xlsx").sheet_names() == ["All-Time"]
    assert WorkbookStore(tmp_path / "R.xlsx").sheet_names() == ["20251129 CW"]


@pytest.fixture
def fake_sota(monkeypatch):
    base = pathlib.Path(__file__).parent / "fixtures"
    activator = (base / "sota_activator.json").read_bytes()
    chaser = (base / "sota_chaser.json").read_bytes()

    class FakeClient:
        def get_bytes(self, url, **kwargs):
            return chaser if "/chaser/" in url else activator

    monkeypatch.setattr("callsigns.providers.sota.HttpClient", lambda: FakeClient())


def test_sota_activator_refresh_then_exports(tmp_path, fake_sota):
    store = tmp_path / "SOTA-Activator.xlsx"
    assert (
        main(
            [
                "refresh",
                "sota-activator",
                "-y",
                "all",
                "-o",
                "cw",
                "--store",
                str(store),
            ]
        )
        == 0
    )
    assert WorkbookStore(store).sheet_names() == ["All-Time CW"]
    assert (
        main(
            [
                "export",
                "sota-activator",
                "dta",
                "-y",
                "all",
                "-o",
                "cw",
                "--store",
                str(store),
                "--out",
                str(tmp_path),
            ]
        )
        == 0
    )
    calls = decode_master_dta((tmp_path / "SOTA-Activator_Calls_CW.dta").read_bytes())
    assert "G4YSS" in calls
    assert "2E0XIS" in calls
    assert "G1OXH" in calls
    assert not any("ANON" in c for c in calls)
    assert not any("#" in c for c in calls)


def test_sota_chaser_has_its_own_store_and_columns(tmp_path, fake_sota):
    store = tmp_path / "SOTA-Chaser.xlsx"
    main(["refresh", "sota-chaser", "-y", "all", "--store", str(store)])
    rows = WorkbookStore(store).read_sheet("All-Time")
    assert "Stations Worked" in rows[0]
    assert "Summits" not in rows[0]


def test_sota_refresh_writes_one_sheet_per_mode(tmp_path, fake_sota):
    store = tmp_path / "S.xlsx"
    main(
        [
            "refresh",
            "sota-activator",
            "-y",
            "all",
            "-o",
            "all,cw",
            "--store",
            str(store),
        ]
    )
    assert sorted(WorkbookStore(store).sheet_names()) == ["All-Time", "All-Time CW"]


def test_sota_xlsx_export_carries_personal_data_columns(tmp_path, fake_sota):
    store = tmp_path / "S.xlsx"
    main(["refresh", "sota-activator", "-y", "all", "--store", str(store)])
    main(
        [
            "export",
            "sota-activator",
            "xlsx",
            "-y",
            "all",
            "--store",
            str(store),
            "--out",
            str(tmp_path),
            "--basename",
            "sota",
        ]
    )
    headers = [c.value for c in load_workbook(tmp_path / "sota.xlsx").active[1]]
    assert "Username" in headers and "Total Points" in headers


def test_all_four_providers_coexist(tmp_path, fake_api, fake_rbn, fake_sota):
    for args in (
        ["refresh", "pota-hunters", "-y", "all", "--store", str(tmp_path / "P.xlsx")],
        [
            "refresh",
            "rbn-cw",
            "-y",
            "20251129",
            "-o",
            "cw",
            "--store",
            str(tmp_path / "R.xlsx"),
        ],
        ["refresh", "sota-activator", "-y", "all", "--store", str(tmp_path / "A.xlsx")],
        ["refresh", "sota-chaser", "-y", "all", "--store", str(tmp_path / "C.xlsx")],
    ):
        assert main(args) == 0
    assert WorkbookStore(tmp_path / "A.xlsx").sheet_names() == ["All-Time"]
    assert WorkbookStore(tmp_path / "C.xlsx").sheet_names() == ["All-Time"]


@pytest.fixture
def fake_contest(monkeypatch):
    base = pathlib.Path(__file__).parent / "fixtures"
    listing = (base / "cq_listing.html").read_text()
    log = (base / "cabrillo_cqww.log").read_text()

    class FakeClient:
        def get_text(self, url, **kwargs):
            return listing

        def get_bytes(self, url, **kwargs):
            return log.encode()

        def content_length(self, url, **kwargs):
            return 100

    client = FakeClient()
    monkeypatch.setattr("callsigns.providers.contest.cq.HttpClient", lambda: client)
    monkeypatch.setattr("callsigns.cache.HttpClient", lambda: client)
    return client


def test_contest_provider_ranks_by_times_worked(tmp_path, fake_contest):
    store = tmp_path / "C.xlsx"
    main(["refresh", "cqww-cw", "-y", "2025", "--store", str(store), "--top-logs", "1"])
    rows = WorkbookStore(store).read_sheet("2025")
    counts = [r["Times Worked"] for r in rows]
    assert counts == sorted(counts, reverse=True)
    assert rows[0]["Callsign"] == "YT6X"


def test_contest_every_export_format(tmp_path, fake_contest):
    store = tmp_path / "C.xlsx"
    main(["refresh", "cqww-cw", "-y", "2025", "--store", str(store), "--top-logs", "1"])
    for fmt, expected in [
        ("dta", "CQWW-CW_Calls_2025.dta"),
        ("scp", "CQWW-CW_Calls_2025.scp"),
        ("lst", "CQWW-CW_Calls_2025.lst"),
    ]:
        assert (
            main(
                [
                    "export",
                    "cqww-cw",
                    fmt,
                    "-y",
                    "2025",
                    "--store",
                    str(store),
                    "--out",
                    str(tmp_path),
                ]
            )
            == 0
        )
        assert (tmp_path / expected).is_file()
    calls = decode_master_dta((tmp_path / "CQWW-CW_Calls_2025.dta").read_bytes())
    assert "YT6X" in calls and "9A1A" in calls
