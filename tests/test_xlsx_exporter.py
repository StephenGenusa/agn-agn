import datetime as dt

from openpyxl import load_workbook

from callsigns.exporters.base import ExportOptions
from callsigns.exporters.xlsx import EXPORT_SHEET_NAME, XlsxExporter
from callsigns.providers.pota import PotaHuntersProvider
from callsigns.store import TABLE_STYLE

PROVIDER = PotaHuntersProvider()

ROWS = [
    {
        "activeCallsign": "K2UPD",
        "numParks": 5529,
        "numQSOs": 6421,
        "qsosCW": 2093,
        "qsosDATA": 642,
        "qsosPHONE": 3686,
    }
]


def options(tmp_path, **kwargs):
    base = {
        "period": "2026",
        "mode": "cw",
        "limit": 500,
        "out_dir": tmp_path,
        "today": dt.date(2026, 8, 9),
    }
    base.update(kwargs)
    return ExportOptions(**base)


def test_filename_follows_the_spec_pattern(tmp_path):
    written = XlsxExporter().write(ROWS, PROVIDER, options(tmp_path))
    assert written[0].name == "POTA-500-CW-2026_2026-08-09.xlsx"


def test_all_period_filename(tmp_path):
    written = XlsxExporter().write(
        ROWS, PROVIDER, options(tmp_path, period="all", mode="all")
    )
    assert written[0].name == "POTA-500-ALL-ALLTIME_2026-08-09.xlsx"


def test_headers_and_values(tmp_path):
    written = XlsxExporter().write(ROWS, PROVIDER, options(tmp_path))
    sheet = load_workbook(written[0]).active
    assert [c.value for c in sheet[1]] == [
        "Callsign",
        "Parks",
        "Total QSOs",
        "Total CW",
        "Total Data",
        "Total Phone",
    ]
    assert [c.value for c in sheet[2]] == ["K2UPD", 5529, 6421, 2093, 642, 3686]


def test_sheet_is_a_blue_banded_table(tmp_path):
    written = XlsxExporter().write(ROWS, PROVIDER, options(tmp_path))
    sheet = load_workbook(written[0]).active
    table = next(iter(sheet.tables.values()))
    assert table.tableStyleInfo.name == TABLE_STYLE
    assert table.tableStyleInfo.showRowStripes is True


def test_no_meta_sheet_in_an_export(tmp_path):
    written = XlsxExporter().write(ROWS, PROVIDER, options(tmp_path))
    assert load_workbook(written[0]).sheetnames == [EXPORT_SHEET_NAME]


def test_default_limit_is_500():
    assert XlsxExporter.default_limit == 500


def test_basename_override(tmp_path):
    written = XlsxExporter().write(ROWS, PROVIDER, options(tmp_path, basename="custom"))
    assert written[0].name == "custom.xlsx"


def test_overwrites_an_existing_export(tmp_path):
    exporter = XlsxExporter()
    exporter.write(ROWS, PROVIDER, options(tmp_path))
    written = exporter.write([], PROVIDER, options(tmp_path))
    sheet = load_workbook(written[0]).active
    assert sheet.max_row == 1
