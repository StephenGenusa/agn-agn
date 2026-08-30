import pytest
from openpyxl import load_workbook

from callsigns.errors import StoreError
from callsigns.providers.base import Column
from callsigns.store import (
    TABLE_STYLE,
    SheetData,
    SheetMeta,
    WorkbookStore,
    table_name_for,
)

COLUMNS = (Column("call", "Callsign", str), Column("n", "Count", int))


def sheet(name, rows, period="all", mode="all"):
    return SheetData(
        name=name,
        columns=COLUMNS,
        rows=rows,
        meta=SheetMeta(
            sheet=name,
            provider="dummy",
            period=period,
            mode=mode,
            rows=len(rows),
            refreshed_utc="2026-08-09T00:00:00+00:00",
            source_url="https://example.test/x",
        ),
    )


def test_table_names_are_sanitised():
    assert table_name_for("All-Time CW") == "T_All_Time_CW"
    assert table_name_for("2026") == "T_2026"


def test_new_store_does_not_exist(tmp_path):
    assert not WorkbookStore(tmp_path / "s.xlsx").exists()


def test_write_then_read_round_trip(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("All-Time", [{"call": "K1ABC", "n": 3}])])
    assert store.exists()
    assert store.read_sheet("All-Time", COLUMNS) == [{"call": "K1ABC", "n": 3}]


def test_sheet_names_exclude_meta(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    assert store.sheet_names() == ["2026"]


def test_no_stray_default_sheet(tmp_path):
    path = tmp_path / "s.xlsx"
    WorkbookStore(path).replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    assert "Sheet" not in load_workbook(path).sheetnames


def test_meta_row_recorded(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    rows = store.meta_rows()
    assert len(rows) == 1
    assert rows[0]["Sheet"] == "2026"
    assert rows[0]["Rows"] == 1
    assert rows[0]["SourceURL"] == "https://example.test/x"


def test_replacing_one_sheet_leaves_others_untouched(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets(
        [
            sheet("2025", [{"call": "W1AW", "n": 9}], period="2025"),
            sheet("2026", [{"call": "K1ABC", "n": 1}], period="2026"),
        ]
    )
    store.replace_sheets([sheet("2026", [{"call": "K9XYZ", "n": 4}], period="2026")])
    assert store.read_sheet("2025", COLUMNS) == [{"call": "W1AW", "n": 9}]
    assert store.read_sheet("2026", COLUMNS) == [{"call": "K9XYZ", "n": 4}]
    assert sorted(store.sheet_names()) == ["2025", "2026"]


def test_meta_updates_only_the_replaced_sheet(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets(
        [
            sheet("2025", [{"call": "W1AW", "n": 9}], period="2025"),
            sheet("2026", [{"call": "K1ABC", "n": 1}], period="2026"),
        ]
    )
    store.replace_sheets(
        [sheet("2026", [{"call": "A", "n": 1}, {"call": "B", "n": 2}], period="2026")]
    )
    by_sheet = {row["Sheet"]: row for row in store.meta_rows()}
    assert by_sheet["2025"]["Rows"] == 1
    assert by_sheet["2026"]["Rows"] == 2


def test_sheet_is_a_styled_table(tmp_path):
    path = tmp_path / "s.xlsx"
    WorkbookStore(path).replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    worksheet = load_workbook(path)["2026"]
    assert worksheet.freeze_panes == "A2"
    tables = list(worksheet.tables.values())
    assert len(tables) == 1
    assert tables[0].tableStyleInfo.name == TABLE_STYLE
    assert tables[0].tableStyleInfo.showRowStripes is True


def test_headers_come_from_columns(tmp_path):
    path = tmp_path / "s.xlsx"
    WorkbookStore(path).replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    worksheet = load_workbook(path)["2026"]
    assert [c.value for c in worksheet[1]] == ["Callsign", "Count"]


def test_empty_sheet_still_writes_headers(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2026", [])])
    assert store.read_sheet("2026", COLUMNS) == []


def test_read_without_columns_uses_headers(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    assert store.read_sheet("2026") == [{"Callsign": "K1ABC", "Count": 1}]


def test_reading_absent_sheet_raises_store_error(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    with pytest.raises(StoreError, match="2025"):
        store.read_sheet("2025")


def test_reading_missing_workbook_raises_store_error(tmp_path):
    with pytest.raises(StoreError, match="no store"):
        WorkbookStore(tmp_path / "absent.xlsx").read_sheet("2026")


def test_creates_parent_directories(tmp_path):
    store = WorkbookStore(tmp_path / "nested" / "deep" / "s.xlsx")
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    assert store.exists()


def test_failed_save_leaves_original_intact(tmp_path, monkeypatch):
    path = tmp_path / "s.xlsx"
    store = WorkbookStore(path)
    store.replace_sheets([sheet("2026", [{"call": "K1ABC", "n": 1}])])
    original = path.read_bytes()

    def boom(*args, **kwargs):
        raise OSError("disk full")

    monkeypatch.setattr("callsigns.store.os.replace", boom)
    with pytest.raises(StoreError):
        store.replace_sheets([sheet("2026", [{"call": "NEW", "n": 2}])])
    assert path.read_bytes() == original
    assert not list(tmp_path.glob("*.tmp"))


def test_drop_sheets_removes_data_and_metadata(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets(
        [
            sheet("2025", [{"call": "W1AW", "n": 9}], period="2025"),
            sheet("2026", [], period="2026"),
        ]
    )
    store.drop_sheets(["2026"])
    assert store.sheet_names() == ["2025"]
    assert [row["Sheet"] for row in store.meta_rows()] == ["2025"]


def test_drop_sheets_ignores_unknown_names(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets([sheet("2025", [{"call": "W1AW", "n": 9}])])
    store.drop_sheets(["nope"])
    assert store.sheet_names() == ["2025"]


def test_drop_sheets_leaves_other_data_intact(tmp_path):
    store = WorkbookStore(tmp_path / "s.xlsx")
    store.replace_sheets(
        [
            sheet("2025", [{"call": "W1AW", "n": 9}], period="2025"),
            sheet("2026", [{"call": "K1ABC", "n": 1}], period="2026"),
        ]
    )
    store.drop_sheets(["2026"])
    assert store.read_sheet("2025", COLUMNS) == [{"call": "W1AW", "n": 9}]
