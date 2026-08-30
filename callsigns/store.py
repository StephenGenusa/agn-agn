"""The workbook of record: sheets of provider rows plus a metadata sheet."""

import os
import pathlib
import re
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from callsigns.errors import StoreError
from callsigns.providers.base import Column

META_SHEET: str = "_meta"
META_HEADERS: tuple[str, ...] = (
    "Sheet",
    "Provider",
    "Period",
    "Mode",
    "Rows",
    "RefreshedUTC",
    "SourceURL",
)
TABLE_STYLE: str = "TableStyleMedium2"
MAX_COLUMN_WIDTH: int = 40


def table_name_for(sheet: str) -> str:
    """Return a valid Excel table name for a sheet.

    Excel table names may not contain spaces or punctuation, so anything
    outside ``[A-Za-z0-9_]`` becomes an underscore, and the name is prefixed
    to guarantee it does not start with a digit.

    Args:
        sheet: The worksheet name.

    Returns:
        A sanitised table name, unique per sheet.
    """
    return "T_" + re.sub(r"[^0-9A-Za-z]+", "_", sheet).strip("_")


@dataclass(frozen=True, slots=True)
class SheetMeta:
    """Provenance for one stored sheet."""

    sheet: str
    provider: str
    period: str
    mode: str
    rows: int
    refreshed_utc: str
    source_url: str


@dataclass(frozen=True, slots=True)
class SheetData:
    """A sheet's columns, rows and provenance, ready to write."""

    name: str
    columns: tuple[Column, ...]
    rows: list[dict[str, object]]
    meta: SheetMeta


def write_data_sheet(worksheet: Worksheet, data: SheetData) -> None:
    """Write headers, rows, a styled table and column widths.

    Shared by the store and by the workbook exporter so both produce
    identically styled sheets.

    Args:
        worksheet: A freshly created, empty worksheet.
        data: The sheet contents to write.
    """
    headers = [column.header for column in data.columns]
    worksheet.append(headers)
    for row in data.rows:
        worksheet.append([row.get(column.key) for column in data.columns])

    last_column = get_column_letter(len(headers))
    last_row = max(len(data.rows) + 1, 2)
    table = Table(
        displayName=table_name_for(data.name),
        ref=f"A1:{last_column}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    worksheet.add_table(table)
    worksheet.freeze_panes = "A2"

    for index, column in enumerate(data.columns, start=1):
        widest = len(column.header)
        for row in data.rows:
            widest = max(widest, len(str(row.get(column.key, ""))))
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = min(widest + 2, MAX_COLUMN_WIDTH)


def atomic_save(book: Workbook, path: pathlib.Path) -> None:
    """Save a workbook so an interrupted write cannot truncate the target.

    Writes a temporary file in the destination directory and then replaces
    the target, which is atomic on a single filesystem.

    Args:
        book: The workbook to save.
        path: Destination path.

    Raises:
        StoreError: The workbook could not be written.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, raw_name = tempfile.mkstemp(
        dir=path.parent, prefix=f".{path.name}.", suffix=".tmp"
    )
    os.close(descriptor)
    temp_path = pathlib.Path(raw_name)
    try:
        book.save(temp_path)
        os.replace(temp_path, path)
    except Exception as exc:
        temp_path.unlink(missing_ok=True)
        raise StoreError(f"cannot write {path}: {exc}") from exc


class WorkbookStore:
    """An Excel workbook holding one sheet per period, plus ``_meta``."""

    def __init__(self, path: pathlib.Path) -> None:
        """Initialise the store.

        Args:
            path: Workbook location. It need not exist yet.
        """
        self._path = path

    @property
    def path(self) -> pathlib.Path:
        """Return the workbook path."""
        return self._path

    def exists(self) -> bool:
        """Return whether the workbook file is present."""
        return self._path.is_file()

    def _load(self, *, read_only: bool = False) -> Workbook:
        """Load the workbook.

        Args:
            read_only: Open in openpyxl's streaming mode, which is far faster
                and far lighter on large sheets but cannot be modified. Used
                for the read paths, where a 90,000-row store would otherwise
                take minutes to parse.

        Returns:
            The open workbook.

        Raises:
            StoreError: The workbook is missing or unreadable.
        """
        if not self.exists():
            raise StoreError(f"no store at {self._path}")
        try:
            return load_workbook(self._path, read_only=read_only, data_only=True)
        except Exception as exc:
            raise StoreError(f"cannot read store {self._path}: {exc}") from exc

    def sheet_names(self) -> list[str]:
        """Return the data sheet names, excluding ``_meta``.

        Returns:
            Sheet names in workbook order.

        Raises:
            StoreError: The workbook is missing or unreadable.
        """
        book = self._load(read_only=True)
        try:
            return [name for name in book.sheetnames if name != META_SHEET]
        finally:
            book.close()

    def read_sheet(
        self, name: str, columns: Sequence[Column] | None = None
    ) -> list[dict[str, object]]:
        """Read one sheet back as row dictionaries.

        Args:
            name: Sheet name.
            columns: The provider's declared columns. When given, rows are
                keyed by column key, so a caller gets back exactly the keys
                it wrote. When omitted, rows are keyed by the display headers
                in row 1.

        Returns:
            Rows in stored order.

        Raises:
            StoreError: The workbook or the sheet is missing.
        """
        book = self._load(read_only=True)
        try:
            if name not in book.sheetnames:
                available = ", ".join(n for n in book.sheetnames if n != META_SHEET)
                raise StoreError(f"store has no sheet {name!r}; available: {available}")
            rows = list(book[name].iter_rows(values_only=True))
        finally:
            book.close()
        if not rows:
            return []
        headers = [str(value) for value in rows[0]]
        if columns is not None:
            by_header = {column.header: column.key for column in columns}
            headers = [by_header.get(header, header) for header in headers]
        return [dict(zip(headers, row, strict=False)) for row in rows[1:]]

    def meta_rows(self) -> list[dict[str, object]]:
        """Return the ``_meta`` sheet as a list of dictionaries.

        Returns:
            One dictionary per recorded sheet, keyed by metadata header.

        Raises:
            StoreError: The workbook is missing or unreadable.
        """
        book = self._load(read_only=True)
        try:
            return self._read_meta(book)
        finally:
            book.close()

    @staticmethod
    def _read_meta(book: Workbook) -> list[dict[str, object]]:
        """Return existing ``_meta`` rows from an open workbook.

        Args:
            book: The open workbook.

        Returns:
            One dictionary per recorded sheet.
        """
        if META_SHEET not in book.sheetnames:
            return []
        rows = list(book[META_SHEET].iter_rows(values_only=True))
        return [dict(zip(META_HEADERS, row, strict=False)) for row in rows[1:]]

    def drop_sheets(self, names: Sequence[str]) -> None:
        """Remove sheets and their metadata rows.

        A period that turns out to hold nothing — a month the source never
        published — leaves an empty sheet behind, which then looks like a
        harvested period that came back blank.

        Args:
            names: Sheet names to remove. Unknown names are ignored.

        Raises:
            StoreError: The workbook could not be read or written.
        """
        book = self._load()
        wanted = set(names)
        removed = [n for n in book.sheetnames if n in wanted and n != META_SHEET]
        for name in removed:
            book.remove(book[name])
        meta = [row for row in self._read_meta(book) if row.get("Sheet") not in wanted]
        if META_SHEET in book.sheetnames:
            book.remove(book[META_SHEET])
        meta_sheet = book.create_sheet(title=META_SHEET)
        meta_sheet.append(list(META_HEADERS))
        for row in meta:
            meta_sheet.append([row.get(header) for header in META_HEADERS])
        meta_sheet.freeze_panes = "A2"
        atomic_save(book, self._path)

    def replace_sheets(self, sheets: Sequence[SheetData]) -> None:
        """Write sheets, replacing same-named ones and leaving others alone.

        Args:
            sheets: The sheets to write.

        Raises:
            StoreError: The workbook could not be read or written.
        """
        if self.exists():
            book = self._load()
        else:
            book = Workbook()
            default = book.active
            if default is not None:
                book.remove(default)

        meta = {str(row["Sheet"]): dict(row) for row in self._read_meta(book)}
        for data in sheets:
            if data.name in book.sheetnames:
                book.remove(book[data.name])
            write_data_sheet(book.create_sheet(title=data.name), data)
            meta[data.meta.sheet] = {
                "Sheet": data.meta.sheet,
                "Provider": data.meta.provider,
                "Period": data.meta.period,
                "Mode": data.meta.mode,
                "Rows": data.meta.rows,
                "RefreshedUTC": data.meta.refreshed_utc,
                "SourceURL": data.meta.source_url,
            }

        if META_SHEET in book.sheetnames:
            book.remove(book[META_SHEET])
        meta_sheet = book.create_sheet(title=META_SHEET)
        meta_sheet.append(list(META_HEADERS))
        for row in meta.values():
            meta_sheet.append([row.get(header) for header in META_HEADERS])
        meta_sheet.freeze_panes = "A2"

        atomic_save(book, self._path)
